import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Configuração da página
st.set_page_config(page_title="PPC - Sistema Fiscal", page_icon="📊", layout="wide")

# --- 1. FUNÇÃO DA MEMÓRIA DE CÁLCULO ---
def gerador_memoria_calculo():
    st.title("📊 Gerador de Memória de Cálculo")
    st.markdown("Arraste a planilha exportada do sistema abaixo:")
    arquivo_upload = st.file_uploader("Selecione o arquivo Excel", type=["xlsx"], key="memoria_up")
    if arquivo_upload:
        st.info("Arquivo recebido. Processando memória de cálculo...")

# --- 2. FUNÇÃO DO RELATÓRIO CONSOLIDADO (ESTRUTURA FINAL REFINADA) ---
def gerador_relatorio_consolidado():
    st.title("📄 Relatório Mensal Consolidado")
    st.markdown("Clique no botão para baixar a estrutura oficial com os ajustes finais de alinhamento e mesclagem.")
    
    if st.button("Gerar Estrutura Oficial"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório Consolidado"
        ws.sheet_view.showGridLines = False

        # --- ESTILOS ---
        azul_ppc = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
        cinza_claro = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        laranja_suave = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        font_branca_bold = Font(color='FFFFFF', bold=True)
        font_preta_bold = Font(color='000000', bold=True)
        borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Largura das colunas
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 60
        ws.column_dimensions['G'].width = 40

        # --- Título Principal (Linha 2) ---
        ws.merge_cells('B2:G2')
        cell_t = ws['B2']
        cell_t.value = "DCTFWeb - Relatório Mensal de Impostos Federais Consolidados"
        cell_t.font = font_branca_bold
        cell_t.fill = azul_ppc
        cell_t.alignment = Alignment(horizontal='center', vertical='center')

        # --- Identificação (Linhas 4 a 7) ---
        dados_id = [
            ("Razão social", "REDCLOUD TECHNOLOGIES BRASIL SERVICOS DIGITAIS LTDA", False),
            ("CNPJ", "47.597.633/0001-92", False),
            ("Período de apuração", "Fevereiro de 2026", True),
            ("Responsável preenchimento", "MARCOS PAULO SANTOS DE OLIVEIRA", False)
        ]

        for i, (label, valor, is_bold) in enumerate(dados_id, 4):
            # Colunas B, C, D (Rótulo) - Centralizado
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
            cell_label = ws.cell(row=i, column=2, value=label)
            cell_label.fill = cinza_claro
            cell_label.font = font_preta_bold
            cell_label.alignment = Alignment(horizontal='center', vertical='center')

            # Colunas E, F, G (Dado) - Alinhado à Esquerda conforme solicitado
            ws.merge_cells(start_row=i, start_column=5, end_row=i, end_column=7)
            cell_valor = ws.cell(row=i, column=5, value=valor)
            cell_valor.fill = cinza_claro
            cell_valor.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            if is_bold:
                cell_valor.font = font_preta_bold

        # --- Cabeçalho da Tabela (Linha 9) ---
        headers = ["Tipo", "Código Retenção", "Valor Retenção", "Descrição do Código da Receita", "", "Observações"]
        for col, text in zip([2, 3, 4, 5, 6, 7], headers):
            cell = ws.cell(row=9, column=col, value=text)
            cell.font = font_branca_bold
            cell.fill = azul_ppc
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = borda_fina
        ws.merge_cells('E9:F9')

        # --- Lista de Impostos (Linha 10 em diante) ---
        impostos = [
            ("INSS", "Folha", "Informação transmitida via eSocial", "Evidência enviada pelo RH"),
            ("IRRF", "0588", "IRRF - Rendimento do Trabalho sem Vínculo Empregatício", ""),
            ("IRRF", "0561", "IRRF - Rendimento do Trabalho Assalariado", ""),
            ("INSS", "1162", "Informação transmitida via EFD REINF - Retenção NFSe", "Memória de cálculo do fiscal"),
            ("IRRF", "1708", "IRRF - Remuneração Serviços Prestados por PJ", ""),
            ("IRRF", "8045", "IRRF - Outros Rendimentos", ""),
            ("IRRF", "3208", "IRRF - Aluguéis e Royalties Pagos a PF", ""),
            ("IRRF", "3280", "IRRF - Rem Serv Prest Associad Coop Trabalho", ""),
            ("CSRF", "5952", "Retenção de Contribuições (CSLL, Cofins e PIS)", ""),
            ("IRRF", "0422", "IRRF - Royalties e Assistência Técnica - Exterior", ""),
            ("PIS", "8109", "PIS - FATURAMENTO - PJ EM GERAL", ""),
            ("COFINS", "2172", "COFINS - FATURAMENTO/PJ EM GERAL", ""),
            ("IRPJ", "2089", "IRPJ - LUCRO PRESUMIDO", ""),
            ("CSLL", "2372", "CSLL - LUCRO PRESUMIDO OU ARBITRADO", "")
        ]

        row_idx = 10
        for tipo, cod, desc, obs in impostos:
            cell_tipo = ws.cell(row=row_idx, column=2, value=tipo)
            cell_tipo.fill = laranja_suave
            cell_tipo.font = font_preta_bold
            
            ws.cell(row=row_idx, column=3, value=cod)
            ws.cell(row=row_idx, column=4, value=0).number_format = 'R$ #,##0.00'
            
            ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=6)
            ws.cell(row=row_idx, column=5, value=desc)
            ws.cell(row=row_idx, column=7, value=obs)

            for c in range(2, 8):
                ws.cell(row=row_idx, column=c).border = borda_fina
                ws.cell(row=row_idx, column=c).alignment = Alignment(horizontal='center', vertical='center')
            row_idx += 1

        # --- Valor Total DARF (Linha Final) ---
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
        cell_l = ws.cell(row=row_idx, column=2, value="Valor Total DARF")
        cell_l.fill = azul_ppc
        cell_l.font = font_branca_bold
        cell_l.border = borda_fina
        cell_l.alignment = Alignment(horizontal='center', vertical='center')

        cell_v = ws.cell(row=row_idx, column=4, value=0)
        cell_v.font = Font(bold=True, color='FF0000') 
        cell_v.border = borda_fina
        cell_v.number_format = 'R$ #,##0.00'
        cell_v.alignment = Alignment(horizontal='center', vertical='center')

        # AJUSTE SOLICITADO: Mesclagem cinza das colunas E, F e G
        ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=7)
        cell_extra_cinza = ws.cell(row=row_idx, column=5)
        cell_extra_cinza.fill = cinza_claro
        
        # Aplicar bordas em todo o bloco mesclado (E, F e G)
        for col_final in range(5, 8):
            ws.cell(row=row_idx, column=col_final).border = borda_fina

        output = BytesIO()
        wb.save(output)
        st.download_button(label="📥 Baixar Estrutura Oficial", data=output.getvalue(), file_name="Relatorio_Consolidado.xlsx")

# --- 3. LÓGICA DE EXIBIÇÃO ---
st.sidebar.title("Navegação")
opcao = st.sidebar.radio("Selecione a ferramenta:", ["Memória de Cálculo", "Relatório Consolidado"])

if opcao == "Memória de Cálculo":
    gerador_memoria_calculo()
else:
    gerador_relatorio_consolidado()

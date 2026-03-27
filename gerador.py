import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image # IMPORTANTE: Adicionado para imagens
import os # IMPORTANTE: Adicionado para localizar o arquivo no servidor

# --- CONFIGURAÇÃO DO CAMINHO DO LOGO ---
# Como o arquivo logo.png está na raiz do repositório no GitHub,
# o Streamlit Cloud o colocará na mesma pasta do script.
# Usamos os.path para garantir que o caminho seja encontrado corretamente.
diretorio_atual = os.path.dirname(os.path.abspath(__file__))
caminho_logo = os.path.join(diretorio_atual, 'logo.png')
# --------------------------------------

# Configuração da página do site
st.set_page_config(page_title="PPC - Gerador", page_icon="📊")

st.title("📊 Gerador de Memória de Cálculo")
st.markdown("Arraste a planilha exportada do sistema abaixo:")

# Área de Upload
arquivo_upload = st.file_uploader("Selecione o arquivo Excel", type=["xlsx"])

def aplicar_estilo_ppc(writer, df_filtrado, colunas_mapeadas, nome_aba, titulo_imposto, razao, cnpj, comp):
    ws = writer.book.create_sheet(nome_aba)
    writer.sheets[nome_aba] = ws
    ws.sheet_view.showGridLines = False 
    ws.column_dimensions['A'].width = 3

    # --- INSERÇÃO DO LOGO ---
    # Tenta carregar e inserir o logo do arquivo caminho_logo
    try:
        if os.path.exists(caminho_logo):
            img = Image(caminho_logo)
            # Redimensiona a imagem para caber na área do cabeçalho
            # Valores aproximados baseados no exemplo, ajuste se necessário
            img.width = 110 
            img.height = 30
            # Adiciona a imagem à célula B2 (canto superior esquerdo)
            ws.add_image(img, 'B2')
        else:
            st.warning(f"Aviso: Arquivo de logo não encontrado em: {caminho_logo}")
    except Exception as e:
        st.error(f"Erro ao inserir o logo: {e}")
    # ------------------------

    fill_azul = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    font_branca = Font(color='FFFFFF', bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')

    # Cabeçalho - Ajustado para alinhar com o exemplo (mesclando a partir da coluna C para deixar espaço pro logo)
    # Define a altura da linha 2 para acomodar o logo
    ws.row_dimensions[2].height = 30

    for row_num, texto in enumerate([f'RAZÃO SOCIAL: {razao}', f'CNPJ: {cnpj}', f'{titulo_imposto} - COMPETÊNCIA {comp}'], 2):
        # Mescla da coluna C (3) até J (10) para deixar espaço para o logo na coluna B
        ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=10)
        cell = ws.cell(row=row_num, column=3) # Texto começa na coluna C
        cell.value = texto
        cell.alignment = align_center
        cell.font = Font(bold=True, size=11) # Fonte um pouco menor para caber melhor

    for col_num, header in enumerate(colunas_mapeadas.values(), 2):
        cell = ws.cell(row=6, column=col_num)
        cell.value = header
        cell.fill = fill_azul
        cell.font = font_branca
        cell.alignment = align_center
        cell.border = thin_border

    if df_filtrado.empty:
        row_msg = 7
        ws.merge_cells(start_row=row_msg, start_column=2, end_row=row_msg, end_column=10)
        cell_msg = ws.cell(row=row_msg, column=2)
        cell_msg.value = "SEM MOVIMENTO"
        cell_msg.font = Font(bold=True)
        cell_msg.alignment = align_center
        for col_idx in range(2, 11):
            ws.cell(row=row_msg, column=col_idx).border = thin_border
    else:
        # --- AJUSTE NO CÓDIGO DE SERVIÇO ---
        # Forçamos a coluna de serviço a ser texto e usar ponto
        for col_orig, col_dest in colunas_mapeadas.items():
            if col_dest == 'Cód. Serviço':
                df_filtrado[col_orig] = df_filtrado[col_orig].astype(str).str.replace(',', '.')

        dados_finais = df_filtrado[list(colunas_mapeadas.keys())].rename(columns=colunas_mapeadas)
        moeda_cols = ['Vlr Contábil', 'Base IRRF', 'Valor IRRF', 'Base CSR', 'Total PCC', 'ISS', 'Valor INSS', 'Base INSS', 'Base ISS']
        
        for r_idx, row in enumerate(dados_finais.values, 7):
            for c_idx, value in enumerate(row, 2):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.border = thin_border
                header_text = list(colunas_mapeadas.values())[c_idx-2]
                
                if header_text in moeda_cols:
                    cell.number_format = 'R$ #,##0.00'
                elif 'Data' in header_text:
                    cell.number_format = 'dd/mm/yyyy'
                elif 'Aliq' in header_text or '%' in header_text:
                    cell.number_format = '0.00%'

        last_row = 6 + len(dados_finais)
        row_total = last_row + 1
        ws.merge_cells(start_row=row_total, start_column=2, end_row=row_total, end_column=6)
        cell_total_label = ws.cell(row=row_total, column=2)
        cell_total_label.value = "TOTAL"
        cell_total_label.font = Font(bold=True)
        cell_total_label.alignment = Alignment(horizontal='right')
        
        for col_idx in range(2, 7):
            ws.cell(row=row_total, column=col_idx).border = thin_border
        
        for col_idx in range(7, len(colunas_mapeadas) + 2):
            ws.cell(row=row_total, column=col_idx).border = thin_border
            header_text = list(colunas_mapeadas.values())[col_idx-2]
            if header_text in moeda_cols:
                col_letter = get_column_letter(col_idx)
                cell_sum = ws.cell(row=row_total, column=col_idx)
                cell_sum.value = f"=SUM({col_letter}7:{col_letter}{last_row})"
                cell_sum.font = Font(bold=True)
                cell_sum.number_format = 'R$ #,##0.00'

    # Ajuste automático de largura, ignorando coluna A e considerando espaço extra
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        if column_letter == 'A': continue
        
        # Para colunas de dados (da linha 6 em diante), o cálculo é automático
        for cell in col:
            if cell.row >= 6 and cell.value: # Pula o cabeçalho no cálculo automático
                length = len(str(cell.value))
                if length > max_length: max_length = length
        
        if max_length > 0:
            ws.column_dimensions[column_letter].width = max_length + 4
        
        # Ajuste manual para colunas do cabeçalho que têm o logo/texto longo e mesclado
        # No exemplo, a coluna C (onde começa o texto) precisa de largura
        # O openpyxl não ajusta automaticamente colunas mescladas.
        if column_letter == 'C':
            ws.column_dimensions['C'].width = 25 # Ajuste manual para caber o texto

# Execução (igual ao original)
if arquivo_upload:
    try:
        df = pd.read_excel(arquivo_upload)
        
        df['ISS_TOTAL'] = df['ISS Dentro do Município'].fillna(0) + df['ISS Fora do Município'].fillna(0)
        df['BASE_ISS_TOTAL'] = df['Base de Cálculo ISS'].fillna(0)
        df['ALIQ_ISS_TOTAL'] = df['% ISS Dentro do Município'].fillna(0) + df['% ISS Fora do Município'].fillna(0)

        razao_cliente = df['Empresa'].iloc[0]
        cnpj_cliente = df['Cnpj Empresa'].iloc[0]
        data_comp = pd.to_datetime(df['Data Competência'].iloc[0])
        comp_formatada = data_comp.strftime('%m.%Y')
        comp_titulo = data_comp.strftime('%m/%Y')

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            m_base = {'Emissão NFe': 'Data Emissão', 'Número NFe': 'Nota Fiscal', 'Serviço Federal': 'Cód. Serviço', 'Prestador': 'Prestador', 'Cnpj/Cpf Prestador': 'CNPJ', 'Valor NFe': 'Vlr Contábil'}
            
            m_1708 = {**m_base, 'Base de Cálculo ISS': 'Base IRRF', '% IRRF': 'Aliq. IRRF', 'Valor IRRF': 'Valor IRRF'}
            m_csrf = {**m_base, 'Base de Cálculo ISS': 'Base CSR', '% CSRF': 'Aliq. CSRF', 'Valor CSRF': 'Total PCC'}
            m_8045 = {**m_base, 'Base de Cálculo ISS': 'Base IRRF', '% IRRF': 'Aliq. IRRF', 'Valor IRRF': 'Valor IRRF'}
            m_iss = {**m_base, 'BASE_ISS_TOTAL': 'Base ISS', 'ALIQ_ISS_TOTAL': 'Aliq. ISS', 'ISS_TOTAL': 'ISS'}
            m_inss = {**m_base, 'Base de Cálculo INSS': 'Base INSS', '% INSS': 'Aliq. INSS', 'Valor INSS': 'Valor INSS'}

            aplicar_estilo_ppc(writer, df[df['DARF IRRF'] == 1708], m_1708, 'IRRF 1708', 'IRRF 1708', razao_cliente, cnpj_cliente, comp_titulo)
            aplicar_estilo_ppc(writer, df[df['DARF CSRF'] == 5952], m_csrf, 'CSRF', 'CSRF', razao_cliente, cnpj_cliente, comp_titulo)
            aplicar_estilo_ppc(writer, df[df['DARF IRRF'] == 8045], m_8045, 'IRRF 8045', 'IRRF 8045', razao_cliente, cnpj_cliente, comp_titulo)
            aplicar_estilo_ppc(writer, df[df['ISS_TOTAL'] > 0], m_iss, 'ISS', 'ISS', razao_cliente, cnpj_cliente, comp_titulo)
            aplicar_estilo_ppc(writer, df[df['Valor INSS'] > 0], m_inss, 'INSS', 'INSS', razao_cliente, cnpj_cliente, comp_titulo)

        st.success(f"✅ Memória de Cálculo de {razao_cliente} pronta!")
        
        st.download_button(
            label="📥 Baixar Memória de Cálculo",
            data=output.getvalue(),
            file_name=f"{razao_cliente} - Memoria de Calculo Retidos {comp_formatada}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"Erro: {e}")
